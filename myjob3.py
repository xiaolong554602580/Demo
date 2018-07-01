#myjob2.py
#-*-coding:utf-8-*-

import json,time,os
from openpyxl import load_workbook
from openpyxl import Workbook



#7月1日加入读取当前目录的.json文件


def listDirFile():
	lists = []
	listDir = os.listdir()

	for lis in listDir:
		if os.path.splitext(lis)[1] == ".json":
			lists.append(lis)
			print("获取到%s"%lis)

	return lists#返回一个目录数组



def save_to_excel(jsonData):#写入专利
	

	try:
		wb = load_workbook("patents.xlsx")
		print("已打开一个patents.xlsx文件")
	except:
		wb = Workbook()
		print("已创建一个patents.xlsx文件")
	
	#wb = load_workbook("patents.xlsx")
	#print("已打开一个patents.xlsx文件")
	ws = wb.active
	rowTitle = ["专利号","专利名称","公开号","专利类型","公开日","申请（专利权）人","发明人","法律状态","简介","主分类号","分类号","地址","代理机构","代理人","国家/省份","摘要"]
	if ws.max_row <= 1: 
		for rowT in range(1,len(rowTitle)+1):
			ws.cell(row=1,column=rowT,value=rowTitle[rowT - 1])

	for j in jsonData:
		with open(j,"r",encoding="utf-8-sig") as jsonF:
			print(">>>>>>>>>>>>>>>>>>>>>>>>打开文件%s"%j)
			jsonN = json.load(jsonF)
			jsonD = jsonN["cubePatentSearchResponse"]["documents"]
		

			for i in range(2,len(jsonD)):
				fs = jsonD[i]["field_values"]
				x = ws.max_row + 1
				ws.cell(row=x,column=1,value=str(fs["id"]))#专利号
				ws.cell(row=x,column=2,value=str(fs["ti"]))#专利名称
				ws.cell(row=x,column=3,value=str(fs["pn"]))#公开号
				ws.cell(row=x,column=4,value=str(fs["type"]))#类型
				#ws.cell(row=i,column=4,value=str(fs["an"]))#申请号
				ws.cell(row=x,column=5,value=str(fs["pd"]))#公开日
				#ws.cell(row=i,column=6,value=str(fs["ad"]))#申请日
				ws.cell(row=x,column=6,value=str(fs["pa"]))#申请（专利权）人
				ws.cell(row=x,column=7,value=str(fs["in"]))#发明人
				ws.cell(row=x,column=8,value=str(fs["ls1"]))#法律状态
				ws.cell(row=x,column=9,value=str(fs["ab"]))#简介
				#ws.cell(row=i,column=11,value=str(fs["apn"]))#重复
				#ws.cell(row=i,column=12,value=str(fs["apd"]))#重复
				ws.cell(row=x,column=10,value=str(fs["ic2"]))#主分类号
				ws.cell(row=x,column=11,value=str(fs["ic1"]))#分类号
				#ws.cell(row=i,column=15,value=str(fs["pr"]))#？
				ws.cell(row=x,column=12,value=str(fs["aa"]))#地址
				ws.cell(row=x,column=13,value=str(fs["agc"]))#代理机构
				ws.cell(row=x,column=14,value=str(fs["agt"]))#代理人
				#ws.cell(row=i,column=19,value=str(fs["cty"]))#国家编号
				#ws.cell(row=i,column=20,value=str(fs["ls1_2"]))#重复
				#ws.cell(row=i,column=21,value=str(fs["ls2_1"]))#无
				#ws.cell(row=i,column=22,value=str(fs["cpa"]))#重复
				#ws.cell(row=i,column=23,value=str(fs["type"]))#类型
				#ws.cell(row=i,column=24,value=str(fs["lsnt"]))#状态1
				#ws.cell(row=i,column=25,value=str(fs["lsn2"]))#状态2
				ws.cell(row=x,column=15,value=str(fs["co"]))#国家/省份
				ws.cell(row=x,column=16,value=str(fs["ac"]))#摘要
				time.sleep(0.05)
				print("**************第%s条专利写入"%(x - 1))

	wb.save("patents.xlsx")#保存文件名
	print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>保存文件,完成！")


if __name__ == '__main__':
	aaa=listDirFile()
	save_to_excel(aaa)
	

